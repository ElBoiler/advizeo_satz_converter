"""
Excel Generator – advizeo Import Templates
==========================================
Generates advizeo-compatible Excel files from parsed BFW DTA ML Satz records.

Supported output templates:
  • create_building_structure  (47 cols)
  • create_tenants             (11 cols)
  • update_tenants             (11 cols)
  • create_devices             (23 cols)
  • update_devices             (14 cols)
"""

from __future__ import annotations
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from parser import LSatz, MSatz, BSatz, DSatz

# ---------------------------------------------------------------------------
# Column definitions (match the actual template column headers exactly)
# ---------------------------------------------------------------------------

BUILDING_STRUCTURE_COLUMNS = [
    "Real estate name",
    "Real estate external identifier",
    "Street and house number",
    "City",
    "Postal code",
    "Country",
    "Yearly due date",
    "Building name",
    "Building external number",
    "Stairway name",
    "Street and house number (stairway)",
    "Estate unit name",
    "Position",
    "Estate unit external identifier",
    "Area [m²]",
    "Heated area [m²]",
    "Warm water area [m²]",
    "Doorplate name",
    "Current Tenant",
    "Moved in date",
    "Moved out date",
    "Email",
    "Phone (primary)",
    "Phone (secondary)",
    "Monthly consumption delivery required",
    "Monthly consumption channel preference",
    "Commercial",
    "Measurement point key",
    "Measurement point",
    "Device type",
    "Consumption groups",
    "Device identifier",
    "Active device counterpart",
    "AES key",
    "Installer email",
    "Start date",
    "End date",
    "Start date due date value",
    "End date due date value",
    "Yearly due date value",
    "Replacement requested?",
    "Device serviced?",
    "HCA coefficient",
    "Calibration year",
    "Heat cost allocator extension used?",
    "Chromium caps used?",
    "Meter module identifier",
]

CREATE_TENANTS_COLUMNS = [
    "Real estate external identifier",
    "Estate unit external identifier",
    "Tenant name",
    "Moved in date",
    "Moved out date",
    "Phone (primary)",
    "Phone (secondary)",
    "Email",
    "Monthly consumption delivery required",
    "Monthly consumption channel preference",
    "Commercial",
]

UPDATE_TENANTS_COLUMNS = CREATE_TENANTS_COLUMNS  # identical structure

CREATE_DEVICES_COLUMNS = [
    "Real estate external identifier",
    "Stairway name",
    "Estate unit external identifier",
    "Measurement point key",
    "Measurement point",
    "Device type",
    "Consumption groups",
    "Device identifier",
    "Active device counterpart",
    "AES key",
    "Installer email",
    "Start date",
    "End date",
    "Start date due date value",
    "End date due date value",
    "Yearly due date value",
    "Replacement requested?",
    "Device serviced?",
    "HCA coefficient",
    "Calibration year",
    "Heat cost allocator extension used?",
    "Chromium caps used?",
    "Meter module identifier",
]

UPDATE_DEVICES_COLUMNS = [
    "Real estate external identifier",
    "Stairway name",
    "Estate unit external identifier",
    "Device identifier",
    "AES key",
    "Installer email",
    "Start date",
    "End date",
    "Replacement requested?",
    "Device serviced?",
    "HCA coefficient",
    "Calibration year",
    "Heat cost allocator extension used?",
    "Chromium caps used?",
]

# ---------------------------------------------------------------------------
# Styling (match advizeo template: dark blue header, white text)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
DATA_FONT   = Font(name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN   = Alignment(vertical="center")


def _apply_header(ws, columns: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(len(col_name) + 2, 16)


def _write_row(ws, row_idx: int, columns: list[str], data: dict[str, Any]) -> None:
    for col_idx, col_name in enumerate(columns, 1):
        val = data.get(col_name, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
        cell.font      = DATA_FONT
        cell.alignment = DATA_ALIGN


def _new_workbook(sheet_name: str, columns: list[str]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    _apply_header(ws, columns)
    return wb, ws


def _to_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Property number → L-Satz lookup helper
# ---------------------------------------------------------------------------

def _prop_lookup(l_saetze: list[LSatz]) -> dict[str, LSatz]:
    """Build a dict: property_number → LSatz (and also external_id → LSatz)."""
    lookup: dict[str, LSatz] = {}
    for l in l_saetze:
        if l.property_number:
            lookup[l.property_number] = l
        if l.external_id and l.external_id != l.property_number:
            lookup[l.external_id] = l
    return lookup


# ---------------------------------------------------------------------------
# Template generators
# ---------------------------------------------------------------------------

def generate_building_structure(parsed: dict) -> bytes:
    """
    Combines L-Satz + M-Satz + B-Satz into the 47-column building structure
    import template.  One row per (tenant × device) combination.
    """
    wb, ws = _new_workbook("Import template", BUILDING_STRUCTURE_COLUMNS)

    l_saetze: list[LSatz] = parsed.get("l_saetze", [])
    m_saetze: list[MSatz] = parsed.get("m_saetze", [])
    b_saetze: list[BSatz] = parsed.get("b_saetze", [])

    prop_map = _prop_lookup(l_saetze)

    # Group M and B by property number
    m_by_prop: dict[str, list[MSatz]] = {}
    for m in m_saetze:
        m_by_prop.setdefault(m.property_number, []).append(m)

    b_by_prop: dict[str, list[BSatz]] = {}
    for b in b_saetze:
        b_by_prop.setdefault(b.property_number, []).append(b)

    row_idx = 2

    for l in l_saetze:
        prop_num = l.property_number
        tenants  = m_by_prop.get(prop_num, [])
        devices  = b_by_prop.get(prop_num, [])

        base: dict[str, Any] = {
            "Real estate name":                  l.real_estate_name,
            "Real estate external identifier":   l.external_id,
            "Street and house number":           l.strasse,
            "City":                              l.stadt,
            "Postal code":                       l.postleitzahl,
            "Country":                           (l.laendercode or "de").lower()[:2],
            "Yearly due date":                   "31.12",
            "Monthly consumption delivery required":  "TRUE",
            "Monthly consumption channel preference": "portal",
            "Commercial": "FALSE",
        }

        # If no tenants, output one property-only row
        if not tenants:
            _write_row(ws, row_idx, BUILDING_STRUCTURE_COLUMNS, base)
            row_idx += 1
            continue

        for m in tenants:
            tenant_base = base.copy()
            unit_id    = m.estate_unit_external_id
            tenant_base.update({
                "Estate unit name":               unit_id or m.apartment_number,
                "Estate unit external identifier": unit_id,
                "Doorplate name":                 m.nutzername1,
                "Current Tenant":                 ("Leerstand" if m.is_vacant
                                                   else m.tenant_name),
                "Moved in date":                  m.einzugsdatum,
                "Moved out date":                 m.auszugsdatum,
            })

            if devices:
                for b in devices:
                    dev_row = tenant_base.copy()
                    dev_row.update({
                        "Device type":       b.device_type_advizeo,
                        "Device identifier": b.geraete_nr,
                        "Start date":        b.abrechnungszeitraum_start,
                        "End date":          b.abrechnungszeitraum_ende,
                        "Replacement requested?": "FALSE",
                        "Device serviced?":       "TRUE",
                    })
                    _write_row(ws, row_idx, BUILDING_STRUCTURE_COLUMNS, dev_row)
                    row_idx += 1
            else:
                _write_row(ws, row_idx, BUILDING_STRUCTURE_COLUMNS, tenant_base)
                row_idx += 1

    # If there are B-Sätze whose property has no L-Satz, add orphan rows
    for b in b_saetze:
        if b.property_number not in prop_map:
            orphan: dict[str, Any] = {
                "Real estate external identifier": b.property_number,
                "Device type":       b.device_type_advizeo,
                "Device identifier": b.geraete_nr,
                "Start date":        b.abrechnungszeitraum_start,
                "End date":          b.abrechnungszeitraum_ende,
                "Replacement requested?": "FALSE",
                "Device serviced?":       "TRUE",
            }
            _write_row(ws, row_idx, BUILDING_STRUCTURE_COLUMNS, orphan)
            row_idx += 1

    return _to_bytes(wb)


def generate_create_tenants(parsed: dict) -> bytes:
    wb, ws = _new_workbook("create_tenants", CREATE_TENANTS_COLUMNS)
    l_saetze: list[LSatz] = parsed.get("l_saetze", [])
    m_saetze: list[MSatz] = parsed.get("m_saetze", [])
    prop_map = _prop_lookup(l_saetze)

    row_idx = 2
    for m in m_saetze:
        l = prop_map.get(m.property_number)
        prop_ext_id = (l.external_id if l else m.property_number) or m.property_number
        row: dict[str, Any] = {
            "Real estate external identifier":  prop_ext_id,
            "Estate unit external identifier":  m.estate_unit_external_id,
            "Tenant name":                      "Leerstand" if m.is_vacant else (m.tenant_name or "Leerstand"),
            "Moved in date":                    m.einzugsdatum,
            "Moved out date":                   m.auszugsdatum,
            "Monthly consumption delivery required":  "TRUE",
            "Monthly consumption channel preference": "portal",
            "Commercial": "FALSE",
        }
        _write_row(ws, row_idx, CREATE_TENANTS_COLUMNS, row)
        row_idx += 1

    return _to_bytes(wb)


def generate_update_tenants(parsed: dict) -> bytes:
    wb, ws = _new_workbook("update_tenants", list(UPDATE_TENANTS_COLUMNS))
    l_saetze: list[LSatz] = parsed.get("l_saetze", [])
    m_saetze: list[MSatz] = parsed.get("m_saetze", [])
    prop_map = _prop_lookup(l_saetze)

    row_idx = 2
    for m in m_saetze:
        l = prop_map.get(m.property_number)
        prop_ext_id = (l.external_id if l else m.property_number) or m.property_number
        row: dict[str, Any] = {
            "Real estate external identifier":  prop_ext_id,
            "Estate unit external identifier":  m.estate_unit_external_id,
            "Tenant name":                      "Leerstand" if m.is_vacant else (m.tenant_name or "Leerstand"),
            "Moved in date":                    m.einzugsdatum,
            "Moved out date":                   m.auszugsdatum,
            "Monthly consumption delivery required":  "TRUE",
            "Monthly consumption channel preference": "portal",
            "Commercial": "FALSE",
        }
        _write_row(ws, row_idx, list(UPDATE_TENANTS_COLUMNS), row)
        row_idx += 1

    return _to_bytes(wb)


def generate_create_devices(parsed: dict) -> bytes:
    wb, ws = _new_workbook("create_devices", CREATE_DEVICES_COLUMNS)
    l_saetze: list[LSatz] = parsed.get("l_saetze", [])
    b_saetze: list[BSatz] = parsed.get("b_saetze", [])
    prop_map = _prop_lookup(l_saetze)

    row_idx = 2
    for b in b_saetze:
        l = prop_map.get(b.property_number)
        prop_ext_id = (l.external_id if l else b.property_number) or b.property_number
        row: dict[str, Any] = {
            "Real estate external identifier": prop_ext_id,
            "Device type":                    b.device_type_advizeo,
            "Device identifier":              b.geraete_nr,
            "Start date":                     b.abrechnungszeitraum_start,
            "End date":                       b.abrechnungszeitraum_ende,
            "Replacement requested?":         "FALSE",
            "Device serviced?":               "TRUE",
        }
        _write_row(ws, row_idx, CREATE_DEVICES_COLUMNS, row)
        row_idx += 1

    return _to_bytes(wb)


def generate_update_devices(parsed: dict) -> bytes:
    wb, ws = _new_workbook("update_devices", UPDATE_DEVICES_COLUMNS)
    l_saetze: list[LSatz] = parsed.get("l_saetze", [])
    b_saetze: list[BSatz] = parsed.get("b_saetze", [])
    prop_map = _prop_lookup(l_saetze)

    row_idx = 2
    for b in b_saetze:
        if not b.geraete_nr:
            continue
        l = prop_map.get(b.property_number)
        prop_ext_id = (l.external_id if l else b.property_number) or b.property_number
        row: dict[str, Any] = {
            "Real estate external identifier": prop_ext_id,
            "Device identifier":              b.geraete_nr,
            "Start date":                     b.abrechnungszeitraum_start,
            "End date":                       b.abrechnungszeitraum_ende,
            "Replacement requested?":         "FALSE",
            "Device serviced?":               "TRUE",
        }
        _write_row(ws, row_idx, UPDATE_DEVICES_COLUMNS, row)
        row_idx += 1

    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# Row-count helpers (for UI display)
# ---------------------------------------------------------------------------

def count_rows(parsed: dict) -> dict[str, int]:
    """Return expected data row counts for each template."""
    l = parsed.get("l_saetze", [])
    m = parsed.get("m_saetze", [])
    b = parsed.get("b_saetze", [])

    # building_structure: one row per (m × b_for_same_prop), or just m if no b
    prop_to_b: dict[str, int] = {}
    for bx in b:
        prop_to_b[bx.property_number] = prop_to_b.get(bx.property_number, 0) + 1

    bs_rows = 0
    for lx in l:
        tenants = sum(1 for mx in m if mx.property_number == lx.property_number)
        devs    = prop_to_b.get(lx.property_number, 0)
        if tenants == 0:
            bs_rows += 1
        elif devs == 0:
            bs_rows += tenants
        else:
            bs_rows += tenants * devs
    # Orphan B rows
    prop_nums = {lx.property_number for lx in l}
    bs_rows += sum(1 for bx in b if bx.property_number not in prop_nums)

    return {
        "building_structure": bs_rows,
        "create_tenants":     len(m),
        "update_tenants":     len(m),
        "create_devices":     len(b),
        "update_devices":     sum(1 for bx in b if bx.geraete_nr),
    }
